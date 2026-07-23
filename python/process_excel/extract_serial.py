# -*- coding: utf-8 -*-
"""
从「中考必背词汇表（1842个）.xlsx」中提取「序号 英文 中文」三列，
按序号顺序输出到新的 Excel 文件。

输出布局：单行表头，下方为 3 个并排的数据块，每块 3 列（序号 英文 中文）：
    序号 | 英文 | 中文 || 序号 | 英文 | 中文 || 序号 | 英文 | 中文
全部条目按序号升序排序后，横排填充：每行连续 3 个条目依次落入 3 个块，
填满一行再换行（序号 1/2/3 在第 1 行，4/5/6 在第 2 行，依此类推）。

源文件结构说明（已分析）：
- 工作表名: 1842，共 1022 行 × 6 列
- 第 1 行: 标题「中考听口词汇总表（带*为2022版课标新增词汇）」
- 采用左右两栏分页排版：
      A列=序号  B列=英文  C列=中文  |  D列=序号  E列=英文  F列=中文
- 表头「序号 英文 中文 序号 英文 中文」在分页处重复出现（约 30 次）
- 序号取值范围 1~1842，无缺失、无重复，合计 1842 个
- 含若干空行分隔符
- 「共用中文」条目：部分短形式与其全称共享同一条中文释义，
  全称行的中文单元格留空，应继承前一个序号（num-1）的中文。
  已知 4 对：app/application(79/80)、gym/gymnasium(679/680)、
  lab/laboratory(839/840)、mathematics/maths(929/930)

处理逻辑：
1. 遍历所有行，分别处理左栏(A,B,C) 与右栏(D,E,F) 两个块
2. 跳过 None、空值、以及表头字符串 '序号'
3. 序号转为整数，以序号为主键去重，按序号升序排序
4. 对中文为空的条目，继承前一个序号（num-1）的中文释义
5. 均分为 3 段，写入 xlsx 的 3 个块（每块 序号/英文/中文）
"""

import argparse
import math
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment

BASE_DIR = Path(__file__).resolve().parent
DEFAULT_SRC = BASE_DIR / "中考必背词汇表（1842个）.xlsx"
DEFAULT_OUT = BASE_DIR / "序号_英文_中文.xlsx"

# 输出分块数量
NUM_BLOCKS = 3

# 源文件中两个块的列索引（0-based）：左栏 (A=序号, B=英文, C=中文)，右栏 (D=序号, E=英文, F=中文)
SRC_BLOCKS = ((0, 1, 2), (3, 4, 5))
HEADER_KEYWORD = "序号"


def _to_int(val) -> int | None:
    """把单元格值转为整数序号；无法转换时返回 None。"""
    if val is None:
        return None
    if isinstance(val, str):
        val = val.strip()
        if val == HEADER_KEYWORD or val == "":
            return None
        try:
            return int(val)
        except ValueError:
            return None
    if isinstance(val, (int, float)):
        return int(val)
    return None


def _to_str(val) -> str:
    """把单元格值转为字符串，None 转为空串。"""
    if val is None:
        return ""
    return str(val).strip()


def extract_entries(xlsx_path: Path) -> list[tuple[int, str, str]]:
    """从 xlsx 中提取所有 (序号, 英文, 中文)，以序号去重并升序排序返回。"""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    if not wb.sheetnames:
        raise ValueError("工作簿中没有任何工作表")
    ws = wb[wb.sheetnames[0]]

    entries: dict[int, tuple[str, str]] = {}
    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        for s_col, e_col, c_col in SRC_BLOCKS:
            if s_col >= len(row):
                continue
            num = _to_int(row[s_col])
            if num is None:
                continue
            english = _to_str(row[e_col]) if e_col < len(row) else ""
            chinese = _to_str(row[c_col]) if c_col < len(row) else ""
            # 以序号为主键去重（后出现者覆盖前者）
            entries[num] = (english, chinese)

    wb.close()

    # 处理「共用中文」情况：部分短形式与其全称（如 app/application、
    # gym/gymnasium、lab/laboratory、mathematics/maths）共享同一条中文释义，
    # 其全称行的中文单元格留空。按序号升序遍历，对中文为空的条目，
    # 继承前一个序号（num-1）的中文释义（升序保证 num-1 已先被处理）。
    for num in sorted(entries):
        en, zh = entries[num]
        if not zh:
            prev = entries.get(num - 1)
            if prev and prev[1]:
                entries[num] = (en, prev[1])

    return [(num, en, zh) for num, (en, zh) in sorted(entries.items(), key=lambda kv: kv[0])]


def write_xlsx(entries: list[tuple[int, str, str]], out_path: Path) -> None:
    """将条目按序号顺序横排写入 xlsx：每行连续 NUM_BLOCKS 个条目，
    依次落入 NUM_BLOCKS 个块（每块 序号/英文/中文），填满一行再换行。"""
    n = len(entries)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "词汇表"

    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    # 第 1 行：表头（NUM_BLOCKS 个块，每块 序号/英文/中文）
    for b in range(NUM_BLOCKS):
        base = b * 3  # 0, 3, 6
        ws.cell(row=1, column=base + 1, value="序号").font = bold
        ws.cell(row=1, column=base + 2, value="英文").font = bold
        ws.cell(row=1, column=base + 3, value="中文").font = bold
        for c in range(base + 1, base + 4):
            ws.cell(row=1, column=c).alignment = center

    # 数据：横排填充——每行连续 NUM_BLOCKS 个条目，依次落入各块
    for i, (num, en, zh) in enumerate(entries):
        b = i % NUM_BLOCKS        # 第几个块（0/1/2）
        r = i // NUM_BLOCKS + 2   # 数据行（表头占第 1 行）
        base = b * 3
        ws.cell(row=r, column=base + 1, value=num).alignment = center
        ws.cell(row=r, column=base + 2, value=en).alignment = left
        ws.cell(row=r, column=base + 3, value=zh).alignment = left

    # 列宽：序号窄，英文/中文宽；每个块重复
    widths = (8, 22, 32)
    for b in range(NUM_BLOCKS):
        base = b * 3
        for offset, w in enumerate(widths):
            ws.column_dimensions[openpyxl.utils.get_column_letter(base + 1 + offset)].width = w

    # 冻结表头
    ws.freeze_panes = "A2"

    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="从 Excel 词汇表中提取「序号 英文 中文」三列并输出到新文件"
    )
    parser.add_argument(
        "input",
        nargs="?",
        default=str(DEFAULT_SRC),
        help=f"输入 Excel 文件路径（默认: {DEFAULT_SRC.name}）",
    )
    parser.add_argument(
        "-o", "--output",
        default=str(DEFAULT_OUT),
        help=f"输出 Excel 文件路径（默认: {DEFAULT_OUT.name}）",
    )
    args = parser.parse_args()

    src_xlsx = Path(args.input)
    out_xlsx = Path(args.output)

    if not src_xlsx.exists():
        raise FileNotFoundError(f"未找到源文件: {src_xlsx}")

    entries = extract_entries(src_xlsx)
    write_xlsx(entries, out_xlsx)

    serials = [n for n, _, _ in entries]
    print(f"提取完成 ✅")
    print(f"  源文件 : {src_xlsx.name}")
    print(f"  输出   : {out_xlsx.name}")
    print(f"  数量   : {len(serials)}")
    if serials:
        print(f"  范围   : {serials[0]} ~ {serials[-1]}")
        # 横排：每行 NUM_BLOCKS 个，块内序号间隔 NUM_BLOCKS
        rows = math.ceil(len(serials) / NUM_BLOCKS)
        print(f"  布局   : 横排，每行 {NUM_BLOCKS} 个，共 {rows} 行")
        # 一致性校验：是否恰好为 1..N 连续序列
        expected = list(range(1, len(serials) + 1))
        if serials == expected:
            print(f"  校验   : 序号 1~{len(serials)} 连续无缺漏 ✓")
        else:
            missing = sorted(set(expected) - set(serials))
            extra = sorted(set(serials) - set(expected))
            print(f"  校验   : ⚠ 存在异常 missing={missing[:10]} extra={extra[:10]}")


if __name__ == "__main__":
    main()
